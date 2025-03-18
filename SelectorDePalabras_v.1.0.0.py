from typing import List
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Tk
import openpyxl

# Lista de palabras comunes en español e inglés que se ignorarán durante el filtrado.
lista_palabras_comunes_espanol_ingles = [word.upper() for word in [
    "de", "la", "que", "el", "en", "y", "a", "los", "del", "se", "las", "por", "un", "para", "con", "no",
    "una", "su", "al", "es", "lo", "como", "más", "pero", "sus", "le", "ya", "o", "este", "sí", "porque",
    "esta", "entre", "cuando", "muy", "sin", "sobre", "también", "me", "hasta", "donde", "quien", "desde",
    "nos", "durante", "siempre", "todo", "todos", "uno", "les", "ni", "contra", "otros", "otro", "etc",
    "poco", "ellos", "e", "tan", "estas", "mucho", "quienes", "nada", "cual", "mío", "mía", "éstos",
    "aquí", "eso", "ésos", "yo", "ante", "i", "me", "my", "myself", "we", "our", "ours", "ourselves",
    "you", "your", "yours", "yourself", "yourselves", "he", "him", "his", "himself", "she", "her", "hers",
    "herself", "it", "its", "itself", "they", "them", "their", "theirs", "themselves", "what", "which",
    "who", "whom", "this", "that", "these", "those", "am", "is", "are", "was", "were", "be", "been", "being",
    "have", "has", "had", "having", "do", "does", "did", "doing", "a", "an", "the", "and", "but", "if", "or",
    "because", "as", "until", "while", "of", "at", "by", "for", "with", "about", "against", "between", "into",
    "through", "during", "before", "after", "above", "below", "to", "from", "up", "down", "in", "out", "on",
    "off", "over", "under", "again", "further", "then", "once", "here", "there", "when", "where", "why", "how",
    "all", "any", "both", "each", "few", "more", "most", "other", "some", "such", "no", "nor", "not", "only",
    "own", "same", "so", "than", "too", "very", "s", "t", "can", "will", "just", "don", "should", "now"
]]

class SelectorDePalabras:
    """
    Clase para seleccionar, modificar y clasificar palabras de un texto.
    Proporciona una interfaz gráfica para revisar palabras filtradas y tomar decisiones sobre ellas.
    """
    def __init__(self):
        """
        Inicializa el SelectorDePalabras.
        Configura el estado inicial de la aplicación y prepara la interfaz de usuario.
        """
        self.ruta_archivo_actual = None  # Nueva variable para ruta actual
        self.ruta_archivo_anterior = None  # Nueva variable para ruta previa
        self.texto = "" # Texto cargado para análisis. Inicialmente vacío.
        self.texto_filtrado_lista = [] # Lista de palabras filtradas del texto. Inicialmente vacía.
        self.strings_elegidos = [] # Lista de palabras que el usuario ha elegido guardar. Inicialmente vacía.
        self.strings_descartados = [] # Lista de palabras que el usuario ha decidido descartar. Inicialmente vacía.
        self.indice_string_actual = 0 # Índice de la palabra actualmente mostrada para revisión. Comienza en 0.
        self.indice_guardado = 0 #Guardaré el mismo indice si abro el mismo archivo
        self.ventana = None # Ventana principal de la interfaz gráfica. Inicialmente no creada.
        self.text_mostrador = None # Widget de texto para mostrar la palabra actual a revisar.
        self.text_previsualizado = None # Widget de texto para previsualizar el contexto de la palabra actual.
        self.estado_modificando = False # Indica si el programa está en modo de modificación de palabra. Inicialmente no.
        self.boton_guardar = None # Botón para guardar la palabra actual.
        self.boton_modificar = None # Botón para modificar la palabra actual.
        self.boton_descartar = None # Botón para descartar la palabra actual.
        self.boton_cancelar_mod = None # Botón para cancelar la modificación (aparece en modo modificación).
        self.boton_abrir_archivo = None # Botón para abrir el diálogo de selección de archivo.
        self.boton_retroceder_palabra = None # Botón para ir a la palabra anterior.
        self.boton_avanzar_palabra = None # Botón para ir a la palabra siguiente.
        self.boton_añadir_delante = None
        self.boton_añadir_atras = None
        self.boton_guardar_mod = None  # Inicializar la variable
        self.boton_cancelar_mod = None  # Inicializar la variable
        self.label_info = None # Etiqueta informativa en la parte superior de la ventana.
        self.etiqueta_progreso = None # Etiqueta para mostrar el progreso de revisión de palabras.
        self.string_original = "" # Almacena la palabra original antes de la modificación.
        self.datos_excel = [] # Lista para almacenar datos que se exportarán a Excel.

    def _filtrar_palabras(self) -> List[str]:
        try:
            if not self.texto:
                raise ValueError("El texto proporcionado está vacío.")

            # Asegurar que el texto sea una cadena
            texto_func = ' '.join(self.texto) if isinstance(self.texto, list) else self.texto
            texto_func = texto_func.upper()  # Convertir a mayúsculas para comparación
            
            patron = r'\(([^)]+)\)|\b[\w./+:-]+\b'
            palabras = []
            palabras_set = set()  # Conjunto para detectar duplicados de forma más eficiente

            for coincidencia in re.finditer(patron, texto_func):
                palabra = coincidencia.group(1) if coincidencia.group(1) else coincidencia.group(0)
                if palabra and len(palabra) > 1 and palabra not in lista_palabras_comunes_espanol_ingles:
                    # Verificar si la palabra ya está en nuestra lista o es parte de otra palabra
                    if palabra not in palabras_set:
                        es_duplicado = False
                        for p in list(palabras_set):
                            if palabra in p or p in palabra:
                                es_duplicado = True
                                break
                        if not es_duplicado:
                            palabras.append(palabra)
                            palabras_set.add(palabra)

            return palabras if palabras else ['No existen palabras para mostrar']

        except Exception as e:
            print(f"Error en filtrar_palabras: {str(e)}")
            return ["NO HAY DATOS"]


    def _abrir_buscador_archivos(self):
        """Abre un diálogo para que el usuario seleccione un archivo de texto."""
        ruta_archivo = filedialog.askopenfilename(
            title="Seleccionar un archivo de texto",
            filetypes=(("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*"))
        )
        if ruta_archivo:
            # Guardar índice y ruta antes de cargar nuevo archivo
            if self.ruta_archivo_actual is not None:
                self.ruta_archivo_anterior = self.ruta_archivo_actual
                self.indice_guardado = self.indice_string_actual
            self._cargar_texto_desde_archivo(ruta_archivo)

    def _cargar_texto_desde_archivo(self, ruta_archivo):
        """Carga el contenido de un archivo y lo convierte en una cadena de texto."""
        try:
            # Verificar si es el mismo archivo que se cargó anteriormente
            if self.ruta_archivo_actual == ruta_archivo:
                respuesta = messagebox.askyesno(
                    "Reanudar",
                    "¿Desea reanudar desde el último índice guardado?",
                    detail=f"Última palabra vista: {self.indice_guardado + 1}"
                )
                self.indice_string_actual = self.indice_guardado if respuesta else 0
            else:
                self.indice_string_actual = 0

            # Guardar la ruta del archivo actual
            self.ruta_archivo_actual = ruta_archivo

            # Leer el contenido del archivo
            with open(ruta_archivo, 'r', encoding='utf-8') as archivo:
                self.texto = archivo.read()  # Esto ya es una cadena (str)
                self.texto_filtrado_lista = self._filtrar_palabras()

                # Verificar si hay palabras para mostrar
                if not self.texto_filtrado_lista or self.texto_filtrado_lista[0] == "NO HAY DATOS":
                    messagebox.showwarning("Advertencia", "No se encontraron palabras para filtrar en este archivo.")
                else:
                    self._mostrar_string_actual()
                    messagebox.showinfo("Archivo Cargado", f"Se ha cargado el archivo: {ruta_archivo}")

                    # Mostrar botones de navegación después de cargar el archivo
                    self.boton_retroceder_palabra.grid(row=4, column=1, pady=5, padx=5, sticky="ew")
                    self.boton_avanzar_palabra.grid(row=4, column=2, pady=5, padx=5, sticky="ew")

        except UnicodeDecodeError:
            messagebox.showerror("Error", "El archivo seleccionado no es un archivo de texto válido.")
        except FileNotFoundError:
            messagebox.showerror("Error", "El archivo seleccionado no se pudo encontrar.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar archivo: {str(e)}")


    def _copiar_archivo(self, ruta_original):
        """Copia el contenido de un archivo a otro."""
        if ruta_original:
            try:
                # Abrir el archivo original en modo lectura
                with open(ruta_original, 'r', encoding='utf-8') as copia:
                    return copia
            except Exception as e:
                print(f"Error al copiar el archivo: {e}")
                return False
        else:
            messagebox.showerror("Error", "Ruta original o nueva no proporcionada.") # Mensaje de error si falla la carga.


    def _mostrar_string_actual(self):
        """Muestra la palabra actual en el widget de texto principal, evitando palabras duplicadas."""
        # Omitir palabras ya procesadas
        while self.indice_string_actual < len(self.texto_filtrado_lista):
            current_word = self.texto_filtrado_lista[self.indice_string_actual]
            if current_word not in self.strings_elegidos and current_word not in self.strings_descartados:
                break
            self.indice_string_actual += 1
        else:
            self._mostrar_resultados_finales()
            return

        if self.indice_string_actual < len(self.texto_filtrado_lista):
            string_actual = self.texto_filtrado_lista[self.indice_string_actual]
            texto_actual = self.text_mostrador.get("1.0", tk.END).strip()

            # Verificar si la palabra actual ya está en el widget de texto principal
            if string_actual != texto_actual and string_actual not in texto_actual.split():
                self.text_mostrador.config(state="normal")
                self.text_mostrador.delete("1.0", tk.END)
                self.text_mostrador.insert(tk.END, string_actual)
                self.text_mostrador.tag_add('center', '1.0', 'end')
                self.text_mostrador.config(state="disabled")

            self.etiqueta_progreso.config(
                text=f"Palabra {self.indice_string_actual + 1} de {len(self.texto_filtrado_lista)}"
            )

            # En _mostrar_string_actual:
            self._mostrar_previsualizacion_texto()



    def _accion_añadir_palabra(self, direccion: int):
        """Añade la palabra anterior o posterior a la palabra en edición, omitiendo palabras procesadas."""
        if not self.estado_modificando:
            return

        if direccion < 0:  # Añadir palabra anterior
            if self.indice_string_actual <= 0:
                return
            self.indice_string_actual -= 1  # Restar 1 para moverte al índice anterior
            indice_palabra = self.indice_string_actual
            
        elif direccion > 0:  # Añadir palabra posterior
            if self.indice_string_actual >= len(self.texto_filtrado_lista) - 1:
                return
            self.indice_string_actual += 1  # Sumar 1 para moverte al índice posterior
            indice_palabra = self.indice_string_actual
        else:
            return  # No se especifica dirección

        # Obtener la palabra a añadir
        palabra = self.texto_filtrado_lista[indice_palabra]

        # Verificar si la palabra ya ha sido procesada
        if palabra in self.strings_elegidos or palabra in self.strings_descartados:
            return  # No añadir palabras ya procesadas

        texto_actual = self.text_mostrador.get("1.0", tk.END).strip()
        
        # Verificar si la palabra ya está en el texto actual
        if palabra.upper() in texto_actual.upper() or texto_actual.upper() in palabra.upper():
            return  # No añadir si la palabra ya está en el texto actual o es parte de él
            
        # Verificar combinaciones que podrían crear duplicados
        if direccion < 0:
            nuevo_texto = palabra + " " + texto_actual
        else:
            nuevo_texto = texto_actual + " " + palabra
            
        # Verificar si hay palabras duplicadas en el nuevo texto
        palabras_nuevo_texto = nuevo_texto.upper().split()
        if len(palabras_nuevo_texto) != len(set(palabras_nuevo_texto)):
            return  # No añadir si crearía palabras duplicadas
            
        self.text_mostrador.delete("1.0", tk.END)
        self.text_mostrador.insert(tk.END, nuevo_texto)

        self._actualizar_previsualizacion_con_resaltado(indice_palabra)

    def _generar_previsualizacion(self, indice_resaltar=None):
        """Función helper para generar la previsualización, evitando palabras duplicadas."""
        # Cálculo de índices (común a ambas funciones)
        palabras_por_lado = 5
        indice_inicio = max(0, self.indice_string_actual - palabras_por_lado)
        indice_fin = min(len(self.texto_filtrado_lista), self.indice_string_actual + palabras_por_lado + 1)

        # Ajustar índices si estamos cerca de los extremos
        if indice_fin - indice_inicio < palabras_por_lado * 2 + 1:
            if self.indice_string_actual < palabras_por_lado:
                indice_fin = min(palabras_por_lado * 2 + 1, len(self.texto_filtrado_lista))
            else:
                indice_inicio = max(0, len(self.texto_filtrado_lista) - (palabras_por_lado * 2 + 1))

        previsualizar = self.texto_filtrado_lista[indice_inicio:indice_fin]

        # Configuración base del widget (común a ambas funciones)
        self.text_previsualizado.config(state="normal")
        self.text_previsualizado.delete("1.0", tk.END)

        # Configurar todos los tags de manera unificada
        tags_config = {
            "resaltado": {"background": "blue", "foreground": "white", "font": ("Consolas", 14, "bold")},
            "guardado": {"foreground": "green"},
            "descartado": {"foreground": "red"},
            "añadido": {"background": "orange", "foreground": "black", "font": ("Consolas", 14, "bold")}
        }
        
        for tag, config in tags_config.items():
            if tag not in self.text_previsualizado.tag_names():
                self.text_previsualizado.tag_configure(tag, **config)

        # Conjunto para rastrear palabras ya añadidas
        palabras_mostradas = set()

        # Insertar palabras con tags
        for i, palabra in enumerate(previsualizar):
            indice_palabra = indice_inicio + i

            # Verificar si la palabra ya ha sido mostrada
            if palabra in palabras_mostradas:
                continue  # Omitir palabras duplicadas
            palabras_mostradas.add(palabra)  # Registrar la palabra como mostrada

            tags = []
            
            if indice_palabra == self.indice_string_actual:
                tags.append("resaltado")
            elif indice_resaltar is not None and indice_palabra == indice_resaltar:
                tags.append("añadido")
            
            if palabra in self.strings_elegidos:
                tags.append("guardado")
            elif palabra in self.strings_descartados:
                tags.append("descartado")
            
            self.text_previsualizado.insert(tk.END, palabra, tuple(tags))
            self.text_previsualizado.insert(tk.END, " ")

        # Finalización común
        self.text_previsualizado.tag_add("center", "1.0", "end")
        self.text_previsualizado.config(state="disabled")
        if not previsualizar:
            return False
        return True

    # Simplificar el método:
    def _actualizar_previsualizacion_con_resaltado(self, indice_resaltar=None):
        """Actualiza la previsualización resaltando la palabra añadida."""
        if not self.texto_filtrado_lista:
            return
        self._generar_previsualizacion(indice_resaltar=indice_resaltar)

    def _mostrar_previsualizacion_texto(self):
        """Muestra la previsualización estándar."""
        if not self.texto_filtrado_lista:
            return
        self._generar_previsualizacion()

    def _refrescar_botones_originales(self):
        """Restaura la visibilidad y posición de los botones principales."""
        self.boton_guardar.grid(row=5, column=1, pady=5, padx=5, sticky="ew")
        self.boton_modificar.grid(row=5, column=0, pady=5, padx=5, sticky="ew")
        self.boton_descartar.grid(row=5, column=2, pady=5, padx=5, sticky="ew")
        self.boton_abrir_archivo.grid(row=4, column=0, pady=5, padx=5, sticky="ew")
        self.boton_retroceder_palabra.grid(row=4, column=1, pady=5, padx=5, sticky="ew")
        self.boton_avanzar_palabra.grid(row=4, column=2, pady=5, padx=5, sticky="ew")

    def _desactivar_botones_modificacion(self):
        self.boton_guardar_mod.grid_remove()
        self.boton_cancelar_mod.grid_remove()
        self.boton_añadir_atras.grid_remove()
        self.boton_añadir_delante.grid_remove()


    def _actualizar_visibilidad_botones(self):
        """Actualiza la visibilidad de los botones según el estado de modificación."""
        if self.estado_modificando:
            # Ocultar botones principales
            self.boton_guardar.grid_remove()
            self.boton_modificar.grid_remove()
            self.boton_descartar.grid_remove()
            self.boton_abrir_archivo.grid_remove()
            self.boton_retroceder_palabra.grid_remove()
            self.boton_avanzar_palabra.grid_remove()
            # Mostrar botones de modificación
            self._activar_botones_modificacion()
        else:
            # Ocultar botones modificacion
            self._desactivar_botones_modificacion()
            # Mostrar botones principales
            self._refrescar_botones_originales()



    def _guardar_string(self, event=None):
        """Guarda la palabra actual en la lista de elegidas y pasa a la siguiente."""
        if self.estado_modificando:
            return # No permite guardar si está en modo modificación.

        if self.indice_string_actual < len(self.texto_filtrado_lista):
            string_a_guardar = self.texto_filtrado_lista[self.indice_string_actual]
            if string_a_guardar not in self.strings_elegidos:
                self.strings_elegidos.append(string_a_guardar) # Añade la palabra a la lista de elegidas.
            self.indice_string_actual += 1 # Incrementa el índice para pasar a la siguiente palabra.
            self._mostrar_string_actual() # Muestra la siguiente palabra.
        return self.strings_elegidos if self.strings_elegidos else 0

    def _descartar_string(self):
        """Descartar la palabra actual, añadiéndola a la lista de descartadas y pasa a la siguiente."""
        if self.estado_modificando:
            return  # No permite descartar si está en modo modificación.

        if self.indice_string_actual < len(self.texto_filtrado_lista):
            string_descartado = self.texto_filtrado_lista[self.indice_string_actual]
            if string_descartado not in self.strings_descartados:
                self.strings_descartados.append(string_descartado)  # Añade la palabra a la lista de descartadas.
            self.indice_string_actual += 1  # Incrementa el índice para pasar a la siguiente palabra.
            self._mostrar_string_actual()  # Muestra la siguiente palabra.

        return self.strings_descartados if self.strings_descartados else 0
    
    def _activar_botones_modificacion(self):
        """Crea y muestra los botones de modificación si no existen, o los reactiva si ya están creados."""
        # Botón para añadir palabra anterior
        if self.boton_añadir_atras is None:
            self.boton_añadir_atras = ttk.Button(self.ventana, text="Añadir Anterior", command=lambda: self._accion_añadir_palabra(-1))
        self.boton_añadir_atras.grid(row=4, column=0, pady=5, padx=5, sticky="ew")
        
        # Botón para añadir palabra posterior
        if self.boton_añadir_delante is None:
            self.boton_añadir_delante = ttk.Button(self.ventana, text="Añadir Posterior", command=lambda: self._accion_añadir_palabra(1))
        self.boton_añadir_delante.grid(row=4, column=2, pady=5, padx=5, sticky="ew")
        
        # Botón para guardar la modificación
        if self.boton_guardar_mod is None:
            self.boton_guardar_mod = ttk.Button(self.ventana, text="Guardar Modificación", command=self._guardar_modificacion)
        self.boton_guardar_mod.grid(row=5, column=0, pady=5, padx=5, sticky="ew", columnspan=2)
        
        # Botón para cancelar la modificación
        if self.boton_cancelar_mod is None:
            self.boton_cancelar_mod = ttk.Button(self.ventana, text="Cancelar Modificación", command=self._cancelar_modificacion)
        self.boton_cancelar_mod.grid(row=5, column=2, pady=5, padx=5, sticky="ew")

    def _activar_modificacion(self):
        """Activa el modo de modificación para la palabra actual."""
        if self.indice_string_actual > len(self.texto_filtrado_lista):
            self.estado_modificando = False
        else:
            self.estado_modificando = True
            # Actualizar visibilidad de los botones
            self._actualizar_visibilidad_botones()
        

        self.string_original = self.texto_filtrado_lista[self.indice_string_actual]
        self.text_mostrador.config(state="normal")
        self.etiqueta_progreso.config(text="Modificando palabra actual...")

    def _guardar_modificacion(self):
        """Guarda la palabra modificada y sale del modo de modificación."""
        nuevo_string = self.text_mostrador.get("1.0", tk.END).strip()

        if nuevo_string:
            # Verificar si hay alguna palabra duplicada en el nuevo string
            palabras = nuevo_string.upper().split()
            if len(palabras) != len(set(palabras)):
                messagebox.showwarning("Advertencia", "No se permiten palabras duplicadas.")
                return

            # Verificar si el nuevo string ya existe en las palabras elegidas o descartadas
            for palabra_existente in self.strings_elegidos + self.strings_descartados:
                if (nuevo_string.upper() == palabra_existente.upper() or 
                    nuevo_string.upper() in palabra_existente.upper() or 
                    palabra_existente.upper() in nuevo_string.upper()):
                    messagebox.showwarning("Advertencia", f"La palabra '{nuevo_string}' ya existe o es similar a '{palabra_existente}'.")
                    return

            # Separar palabras individuales si contiene espacios
            if " " in nuevo_string:
                palabras_individuales = nuevo_string.split()
                # Eliminar la palabra original de la lista
                if self.string_original in self.texto_filtrado_lista:
                    self.texto_filtrado_lista.remove(self.string_original)

                # Agregar cada palabra individual a la lista de palabras filtradas
                for palabra in palabras_individuales:
                    if palabra not in self.texto_filtrado_lista and palabra not in self.strings_elegidos and palabra not in self.strings_descartados:
                        self.texto_filtrado_lista.append(palabra)
                        self.strings_elegidos.append(palabra)
            else:
                # Es una sola palabra, reemplazar la original
                self.texto_filtrado_lista[self.indice_string_actual] = nuevo_string
                if nuevo_string not in self.strings_elegidos:
                    self.strings_elegidos.append(nuevo_string)

            self.indice_string_actual += 1

            # Salir del modo modificación
            self.estado_modificando = False
            self.text_mostrador.config(state="disabled")

            # Actualizar interfaz
            self.etiqueta_progreso.config(text=f"Palabra {self.indice_string_actual + 1} de {len(self.texto_filtrado_lista)}")
            self._mostrar_string_actual()
            # Actualizar visibilidad de los botones
            self._actualizar_visibilidad_botones()
            return nuevo_string
        else:
            # Si no hay modificación, simplemente cancelar
            self._cancelar_modificacion()
            # Actualizar visibilidad de los botones
            self._actualizar_visibilidad_botones()

    def _cancelar_modificacion(self):
        """Cancela la modificación, restaura la palabra original y sale del modo de modificación."""
        self.estado_modificando = False

        # Actualizar visibilidad de los botones
        self._actualizar_visibilidad_botones()

        # Restaurar palabra original en el mostrador
        self.text_mostrador.config(state="normal")
        self.text_mostrador.delete("1.0", tk.END)
        if self.string_original not in self.text_mostrador.get("1.0", tk.END):
            self.text_mostrador.insert(tk.END, self.string_original)
        self.text_mostrador.tag_add('center', '1.0', 'end')
        self.text_mostrador.config(state="disabled")
        
        self.etiqueta_progreso.config(text=f"Palabra {self.indice_string_actual + 1} de {len(self.texto_filtrado_lista)}")
        self._mostrar_previsualizacion_texto()

    def _avanzar_retroceder(self, direccion):
        """Avanza o retrocede en la lista de palabras filtradas, omitiendo procesadas."""
        if self.estado_modificando:
            return
        
        step = 1 if direccion > 0 else -1
        nuevo_indice = self.indice_string_actual + step
        
        while 0 <= nuevo_indice < len(self.texto_filtrado_lista):
            current_word = self.texto_filtrado_lista[nuevo_indice]
            if current_word not in self.strings_elegidos and current_word not in self.strings_descartados:
                self.indice_string_actual = nuevo_indice
                self._mostrar_string_actual()
                return
            nuevo_indice += step



    def _mostrar_resultados_finales(self):
        """Muestra la interfaz final con los resultados de palabras elegidas y descartadas."""
        for widget in self.ventana.winfo_children():
            widget.destroy() # Limpia todos los widgets de la ventana principal.

        notebook = ttk.Notebook(self.ventana) # Crea un notebook (pestañas) para organizar los resultados.
        notebook.pack(fill="both", expand=True, padx=10, pady=10) # Empaqueta el notebook para que se expanda en la ventana.

        tab_elegidas = ttk.Frame(notebook) # Crea una pestaña para palabras elegidas.
        notebook.add(tab_elegidas, text="Palabras Elegidas") # Añade la pestaña al notebook con su título.

        tab_descartadas = ttk.Frame(notebook) # Crea una pestaña para palabras descartadas.
        notebook.add(tab_descartadas, text="Palabras Descartadas") # Añade la pestaña al notebook con su título.

        self._crear_lista_resultados(tab_elegidas, list(self.strings_elegidos)) # Crea la lista de resultados en la pestaña de elegidas.
        self._crear_lista_resultados(tab_descartadas, self.strings_descartados) # Crea la lista de resultados en la pestaña de descartadas.

        frame_botones = ttk.Frame(self.ventana) # Crea un frame para los botones de la parte inferior.
        frame_botones.pack(pady=10) # Empaqueta el frame de botones.

        ttk.Button(
            frame_botones,
            text="Exportar Resultados", # Texto del botón.
            command=self._exportar_resultados # Comando para exportar al hacer clic.
        ).pack(side="left", padx=5) # Posiciona el botón de exportar.

        ttk.Button(
            frame_botones,
            text="Cerrar", # Texto del botón.
            command=self.ventana.destroy # Comando para cerrar la ventana al hacer clic.
        ).pack(side="left", padx=5) # Posiciona el botón de cerrar.

    def _crear_lista_resultados(self, contenedor, lista_palabras):
        """Crea una lista desplazable (listbox) para mostrar las palabras en una pestaña de resultados."""
        frame = ttk.Frame(contenedor) # Crea un frame dentro del contenedor (pestaña).
        frame.pack(fill="both", expand=True, padx=10, pady=10) # Empaqueta el frame para que se expanda.

        scrollbar = ttk.Scrollbar(frame) # Crea una barra de desplazamiento vertical.
        scrollbar.pack(side="right", fill="y") # Posiciona la barra de desplazamiento a la derecha.

        listbox = tk.Listbox(
            frame,
            font=("Consolas", 10), # Fuente del texto en la lista.
            yscrollcommand=scrollbar.set # Asocia la barra de desplazamiento con el listbox.
        )
        listbox.pack(side="left", fill="both", expand=True) # Posiciona el listbox para que se expanda.
        scrollbar.config(command=listbox.yview) # Configura la barra de desplazamiento para controlar el listbox.

        for palabra in sorted(lista_palabras): # Inserta cada palabra en el listbox, ordenadas alfabéticamente.
            listbox.insert(tk.END, palabra)

        ttk.Label(
            contenedor,
            text=f"Total: {len(lista_palabras)} palabras" # Muestra el total de palabras en la lista.
        ).pack(pady=5) # Empaqueta la etiqueta con el total de palabras.

    def _exportar_resultados(self):
        """Exporta las palabras elegidas y descartadas a un archivo de texto."""
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".txt", # Extensión de archivo por defecto.
                filetypes=[("Archivos de texto", "*.txt"), ("Todos los archivos", "*.*")], # Filtros de tipo de archivo.
                title="Guardar resultados" # Título del diálogo para guardar.
            )

            if not archivo: # Si el usuario cancela el diálogo de guardar, sale de la función.
                return

            with open(archivo, "w", encoding="utf-8") as f: # Abre el archivo para escritura en modo texto UTF-8.
                f.write("=== PALABRAS ELEGIDAS ===\n") # Escribe el encabezado para palabras elegidas.
                for palabra in sorted(list(self.strings_elegidos)): # Escribe cada palabra elegida, ordenadas alfabéticamente.
                    f.write(f"{palabra}\n")

                f.write("\n\n=== PALABRAS DESCARTADAS ===\n") # Escribe el encabezado para palabras descartadas.
                for palabra in sorted(self.strings_descartados): # Escribe cada palabra descartada, ordenadas alfabéticamente.
                    f.write(f"{palabra}\n")

            messagebox.showinfo("Exportación Exitosa", f"Resultados guardados en {archivo}") # Mensaje de éxito al exportar.

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar resultados: {str(e)}") # Mensaje de error si falla la exportación.

    def agregar_dato(self, *valores):
        """Agrega datos para exportar a Excel (función no utilizada directamente en la interfaz en este código)."""
        if valores not in self.datos_excel: # Evita duplicados.
            self.datos_excel.append(valores) # Añade los valores a la lista de datos para Excel.

    def generar_excel(self):
        """Genera un archivo Excel con los resultados de palabras elegidas y descartadas."""
        root = Tk()
        root.withdraw()  # Oculta la ventana principal de Tkinter para el diálogo de guardar archivo.
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",  # Extensión de archivo por defecto para Excel.
            filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]  # Filtros de tipo de archivo para Excel.
        )

        if not ruta_archivo:  # Si el usuario cancela el diálogo de guardar, sale de la función.
            return False

        workbook = openpyxl.Workbook()  # Crea un nuevo libro de Excel.
        hoja = workbook.active  # Selecciona la hoja activa del libro.
        hoja.title = "Palabras Seleccionadas"  # Nombre de la hoja

        # Procesar las palabras elegidas para separar palabras individuales si es necesario
        palabras_elegidas_procesadas = []
        for palabra in self.strings_elegidos:
            if " " in palabra:
                # Si la palabra contiene espacios, separarla en palabras individuales
                palabras_individuales = palabra.split()
                for p_individual in palabras_individuales:
                    if p_individual not in palabras_elegidas_procesadas:
                        palabras_elegidas_procesadas.append(p_individual)
            else:
                # Si es una palabra única, agregarla directamente
                if palabra not in palabras_elegidas_procesadas:
                    palabras_elegidas_procesadas.append(palabra)

        # Procesar las palabras descartadas de la misma manera
        palabras_descartadas_procesadas = []
        for palabra in self.strings_descartados:
            if " " in palabra:
                palabras_individuales = palabra.split()
                for p_individual in palabras_individuales:
                    if p_individual not in palabras_descartadas_procesadas:
                        palabras_descartadas_procesadas.append(p_individual)
            else:
                if palabra not in palabras_descartadas_procesadas:
                    palabras_descartadas_procesadas.append(palabra)

        # Crear cabeceras en el Excel
        hoja.cell(row=1, column=1, value="Palabras Elegidas")
        hoja.cell(row=1, column=2, value="Palabras Descartadas")
        
        # Ordenar las palabras alfabéticamente
        palabras_elegidas_procesadas.sort()
        palabras_descartadas_procesadas.sort()
        
        # Llenar el Excel con las palabras procesadas
        max_len = max(len(palabras_elegidas_procesadas), len(palabras_descartadas_procesadas))
        for i in range(max_len):
            # Palabra elegida
            if i < len(palabras_elegidas_procesadas):
                hoja.cell(row=i+2, column=1, value=palabras_elegidas_procesadas[i])
            
            # Palabra descartada
            if i < len(palabras_descartadas_procesadas):
                hoja.cell(row=i+2, column=2, value=palabras_descartadas_procesadas[i])

        try:
            workbook.save(ruta_archivo)  # Guarda el libro de Excel en la ruta especificada.
            return True  # Indica que la exportación a Excel fue exitosa.
        except Exception as e:
            print(f"Error al guardar el archivo: {e}")  # Imprime un mensaje de error en la consola si falla la exportación.
            messagebox.showerror("Error", f"No se pudo guardar el archivo Excel: {e}")
            return False

    def on_closing(self):
        # Guardar índice y ruta actual antes de cerrar
        if self.ruta_archivo_actual:
            self.indice_guardado = self.indice_string_actual
            self.ruta_archivo_anterior = self.ruta_archivo_actual
            
        if self.strings_elegidos or self.strings_descartados:
            if messagebox.askyesno("Salir", "¿Desea generar un archivo Excel con los datos seleccionados?"):
                if self.generar_excel():
                    messagebox.showinfo("Éxito", "Archivo Excel generado correctamente.")
        
        self.ventana.destroy()
    
    def run(self):
        self.ventana = tk.Tk()
        self.ventana.title("Selector de Palabras")
        self.ventana.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Establecer un tamaño mediano-grande (800x600)
        self.ventana.geometry("800x600")

        # Centrar la ventana en el monitor
        self.ventana.update_idletasks()
        ancho_ventana = self.ventana.winfo_width()
        alto_ventana = self.ventana.winfo_height()
        ancho_pantalla = self.ventana.winfo_screenwidth()
        alto_pantalla = self.ventana.winfo_screenheight()
        x = (ancho_pantalla // 2) - (ancho_ventana // 2)
        y = (alto_pantalla // 2) - (alto_ventana // 2)
        self.ventana.geometry(f"+{x}+{y}")

        # Configuración de la interfaz gráfica
        self.label_info = ttk.Label(self.ventana, text="Bienvenido al Selector de Palabras.\nCarga un archivo de texto para comenzar.")
        self.label_info.grid(row=0, column=0, columnspan=3, pady=10)

        self.etiqueta_progreso = ttk.Label(self.ventana, text="Carga un archivo para comenzar")
        self.etiqueta_progreso.grid(row=1, column=0, columnspan=3, pady=5)

        # Crear un frame para los widgets de texto
        frame_texto = ttk.Frame(self.ventana)
        frame_texto.grid(row=2, column=0, columnspan=3, sticky="nsew", padx=10, pady=5)
        self.ventana.grid_rowconfigure(2, weight=1)
        
        self.text_previsualizado = tk.Text(frame_texto, height=3, state="disabled", wrap="word", font=("Consolas", 14))
        self.text_previsualizado.tag_configure("center", justify='center')
        self.text_previsualizado.pack(fill="both", expand=True, pady=5)

        self.text_mostrador = tk.Text(frame_texto, height=2, state="disabled", wrap="word", font=("Arial", 16))
        self.text_mostrador.tag_configure("center", justify='center')
        self.text_mostrador.pack(fill="both", expand=True, pady=5)

        # Crear frames para botones
        frame_botones_nav = ttk.Frame(self.ventana)
        frame_botones_nav.grid(row=3, column=0, columnspan=3, sticky="ew", pady=5)
        
        frame_botones_accion = ttk.Frame(self.ventana)
        frame_botones_accion.grid(row=4, column=0, columnspan=3, sticky="ew", pady=5)
        

        self.boton_abrir_archivo = ttk.Button(frame_botones_nav, text="Abrir Documento", command=self._abrir_buscador_archivos)
        self.boton_abrir_archivo.grid(row=0, column=0, padx=5, sticky="ew")
        
        self.boton_retroceder_palabra = ttk.Button(frame_botones_nav, text="Anterior", command=lambda: self._avanzar_retroceder(-1))
        self.boton_retroceder_palabra.grid(row=0, column=1, padx=5, sticky="ew")
        self.boton_retroceder_palabra.grid_remove()  # Ocultar inicialmente
        
        self.boton_avanzar_palabra = ttk.Button(frame_botones_nav, text="Siguiente", command=lambda: self._avanzar_retroceder(1))
        self.boton_avanzar_palabra.grid(row=0, column=2, padx=5, sticky="ew")
        self.boton_avanzar_palabra.grid_remove()  # Ocultar inicialmente
        
        # Hacer lo mismo con los botones de acción
        self.boton_modificar = ttk.Button(frame_botones_accion, text="Modificar", command=self._activar_modificacion)
        self.boton_modificar.grid(row=0, column=0, padx=5, sticky="ew")
        
        self.boton_guardar = ttk.Button(frame_botones_accion, text="Guardar", command=self._guardar_string)
        self.boton_guardar.grid(row=0, column=1, padx=5, sticky="ew")
        
        self.boton_descartar = ttk.Button(frame_botones_accion, text="Descartar", command=self._descartar_string)
        self.boton_descartar.grid(row=0, column=2, padx=5, sticky="ew")
        
        # Configurar columnas para que sean de igual tamaño
        frame_botones_nav.columnconfigure((0, 1, 2), weight=1)
        frame_botones_accion.columnconfigure((0, 1, 2), weight=1)

        # Configuración para que la ventana sea responsive
        self.ventana.columnconfigure(0, weight=1)
        self.ventana.rowconfigure(2, weight=1)
        self.ventana.mainloop()

if __name__ == "__main__":
    selector = SelectorDePalabras() # Crea una instancia del SelectorDePalabras.
    selector.run() # Ejecuta la interfaz gráfica de la instancia creada.